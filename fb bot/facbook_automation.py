from cmath import e
from types import NoneType
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from bs4 import BeautifulSoup
import openpyxl
import requests

def facebook():
    wb_post = openpyxl.Workbook()
    all_post_sheet = wb_post.active

    all_post_sheet.cell(row = 1, column = 1).value = "Sr NO."
    all_post_sheet.cell(row = 1, column = 2).value = "Post Caption"
    all_post_sheet.cell(row = 1, column = 3).value = "Post Image Link"
    all_post_sheet.cell(row = 1, column = 4).value = "Post website Link"
    all_post_sheet.cell(row = 1, column = 5).value = "website Title"

    #chrome userdata folder directory address 
    userdatadir = 'C:\\test\\1_22\\User Data'
    chromeOptions = webdriver.ChromeOptions() 
    chromeOptions.add_argument(f"--user-data-dir={userdatadir}")

    #chrome profile number
    chromeOptions.add_argument('--profile-directory=Profile 1') 
    driver = webdriver.Chrome(options=chromeOptions)
    driver.get('https://www.facebook.com/noviba')
    time.sleep(3)
    for i in range(0,120):
        time.sleep(2)
        driver.execute_script("scrollBy(0,+1000);")
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    #print(soup.prettify())

    main_timeline_div = soup.find('div',attrs={'data-pagelet':'ProfileTimeline'})
    try:
        all_posts = main_timeline_div.find_all('div', attrs={'class','j83agx80 l9j0dhe7 k4urcfbm'})
    except e:
        print(e)
    count = 1
    row = 2
    web_link_list = []
    for post in all_posts:
        time_slot = post.find('a',attrs={'class':'oajrlxb2 g5ia77u1 qu0x051f esr5mh6w e9989ue4 r7d6kgcz rq0escxv nhd2j8a9 nc684nl6 p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso i1ao9s8h esuyzwwr f1sip0of lzcic4wl gmql0nx0 gpro0wi8 b1v8xokw'})
        if time_slot:
            text_time = time_slot['aria-label']
            #print(text_time)
        else:
            break
        web_link = ""
        link_text = ""
        post_caption_text = ""
        if "m" in str(text_time) or "h" in str(text_time):
            post_caption = post.find('div',attrs={'class','kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x c1et5uql ii04i59q'})
            if post_caption == "" or post_caption == "None" or post_caption == None or post_caption is NoneType:
                pass
            else:
                post_caption_text = post_caption.text
            #print(post_caption_text)
            post_image_link = post.find('img',attrs={'class','i09qtzwb n7fi1qx3 datstx6m pmk7jnqg j9ispegn kr520xx4 k4urcfbm bixrwtb6'})
            if post_image_link:
                link_text = post_image_link['src']
                #print(link_text)
            else:
                pass
            novi_ba_link = post.find('a',attrs={'class','oajrlxb2 g5ia77u1 qu0x051f esr5mh6w e9989ue4 r7d6kgcz rq0escxv nhd2j8a9 p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso i1ao9s8h esuyzwwr f1sip0of lzcic4wl gmql0nx0 a8c37x1j p8dawk7l'})
            if novi_ba_link:
                novi_link_text = novi_ba_link['href']
                #print(f"flink: {novi_link_text}")
                web_link = novi_link_text
            else:
                novi_ba = post.find('a',attrs={'class','oajrlxb2 g5ia77u1 qu0x051f esr5mh6w e9989ue4 r7d6kgcz rq0escxv nhd2j8a9 nc684nl6 p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso i1ao9s8h esuyzwwr f1sip0of lzcic4wl i09qtzwb n7fi1qx3 pmk7jnqg j9ispegn kr520xx4 dwzzwef6'})
                if novi_ba == "" or novi_ba == None or novi_ba is NoneType or novi_ba == "None":
                    pass
                else:
                    novi_ba_text_novi = novi_ba['href']
                    #print(f"slink: {novi_ba_text_novi}")
                    web_link = novi_ba_text_novi
            web_link_list.append(web_link)
            all_post_sheet.cell(row = row, column = 1).value = count
            all_post_sheet.cell(row = row, column = 2).value = post_caption_text
            all_post_sheet.cell(row = row, column = 3).value = link_text
            #all_post_sheet.cell(row = row, column = 4).value = web_link
            count += 1
            row += 1
        else:
            break
    row_again = 2
    for link in web_link_list:
        if "novi.ba" in link:
            driver.get(link)
            time.sleep(1)
            for i in range(0,3):
                time.sleep(1)
                driver.execute_script("scrollBy(0,+800);")
            webhtml = driver.page_source
            soup = BeautifulSoup(webhtml, 'html.parser')
            #print(soup.prettify())
            title_check1 = soup.find('header', attrs={'class','td-post-title'})
            title_check2 = soup.find('div', attrs={'class','head_article'})
            title_check3 = soup.find('h1', attrs={'class','article-menu-trigger'})
            if title_check1:
                web_title = title_check1.find('h1', attrs={'class','entry-title'})
                website_title = web_title.text
                #print(f"Website Tile: {website_title}")
                all_post_sheet.cell(row = row_again, column = 5).value = website_title
                get_url = driver.current_url
                all_post_sheet.cell(row = row_again, column = 4).value = get_url
                row_again +=1
            elif title_check2:
                web_title = title_check2.find('h1')
                website_title = web_title.text
                #print(f"Website Tile: {website_title}")
                all_post_sheet.cell(row = row_again, column = 5).value = website_title
                get_url = driver.current_url
                all_post_sheet.cell(row = row_again, column = 4).value = get_url
                row_again +=1
            elif title_check3:
                website_title = title_check3.text
                #print(f"Website Tile: {website_title}")
                all_post_sheet.cell(row = row_again, column = 5).value = website_title
                get_url = driver.current_url
                all_post_sheet.cell(row = row_again, column = 4).value = get_url
                row_again +=1
            else:
                pass
        else:
            pass

    driver.close()
    driver.quit()
    wb_post.save("posts.xlsx")

def download_images():
    book = openpyxl.load_workbook('posts.xlsx')
    sheet = book.active
    row_count = sheet.max_row
    #print(row_count)
    for row in range(2, int(row_count)+1):
        val = sheet.cell(row = row, column = 3).value
        img_data = requests.get(val).content
        with open(f'post_image{row}.jpg', 'wb') as handler:
            handler.write(img_data)



def main():
    facebook()
    download_images()

if __name__ == "__main__":
    main()